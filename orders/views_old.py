from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login
from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic.edit import CreateView
from django.views.generic import TemplateView, ListView, DetailView, UpdateView
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.mail import EmailMessage
from django.conf import settings
from django.template.loader import render_to_string
from django.utils import timezone
from django.core.cache import cache
from django.db.models import Q
from datetime import datetime
from .models import Factory, Order, Notification, NotificationSettings, Country, OrderConfirmation, OrderAuditLog
from .forms import OrderForm, InvoiceUploadForm, CountryForm, FactoryForm, NotificationSettingsForm, NotificationFilterForm
from .file_preview import generate_file_preview
from .tasks import send_order_notification
from .analytics import get_analytics_data
import os


class HomeView(TemplateView):
    template_name = 'orders/home.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user'] = self.request.user
        
        if self.request.user.is_authenticated:
            # Кэшируем статистику для пользователя на 5 минут
            cache_key = f"user_stats_{self.request.user.id}"
            cached_stats = cache.get(cache_key)
            
            if cached_stats is None:
                # Статистика для аутентифицированного пользователя - оптимизированный запрос
                from django.db.models import Case, When, IntegerField, Q, Count
                from datetime import timedelta
                
                now = timezone.now()
                week_ago = now - timedelta(days=7)
                
                # Один запрос для всех статистик
                user_orders = Order.objects.filter(employee=self.request.user)
                
                # Получаем статистику по статусам одним запросом
                status_stats = user_orders.aggregate(
                    total_orders=Count('id'),
                    uploaded_orders=Count('id', filter=Q(status='uploaded')),
                    sent_orders=Count('id', filter=Q(status='sent')),
                    completed_orders=Count('id', filter=Q(status='completed')),
                    overdue_orders=Count('id', filter=Q(
                        Q(status='uploaded', uploaded_at__lte=week_ago) |
                        Q(status='sent', sent_at__lte=week_ago)
                    ))
                )
                
                # Получаем последние заказы с оптимизацией
                recent_orders = list(user_orders.select_related('factory').filter(
                    id__isnull=False
                ).order_by('-uploaded_at')[:5].values(
                    'id', 'title', 'status', 'uploaded_at', 'factory__name'
                ))
                
                # Дополнительные данные для виджетов
                # Статистика по фабрикам
                factory_stats = Factory.objects.filter(
                    order__employee=self.request.user
                ).annotate(
                    total_orders=Count('order'),
                    active_orders=Count('order', filter=Q(order__status__in=['uploaded', 'sent', 'invoice_received']))
                ).order_by('-total_orders')[:5]
                
                # Статистика по месяцам (последние 6 месяцев)
                six_months_ago = datetime.now() - timedelta(days=180)
                monthly_stats = Order.objects.filter(
                    employee=self.request.user,
                    uploaded_at__gte=six_months_ago
                ).extra(
                    select={'month': "DATE_TRUNC('month', uploaded_at)"}
                ).values('month').annotate(
                    count=Count('id')
                ).order_by('month')
                
                # Заказы, требующие внимания
                urgent_orders = Order.objects.filter(
                    employee=self.request.user,
                    status='uploaded',
                    uploaded_at__lt=datetime.now() - timedelta(days=7)
                ).select_related('factory').order_by('uploaded_at')[:3]
                
                # Статистика по странам
                country_stats = Country.objects.filter(
                    factory__order__employee=self.request.user
                ).annotate(
                    total_orders=Count('factory__order'),
                    total_factories=Count('factory', distinct=True)
                ).order_by('-total_orders')[:5]
                
                stats = {
                    'total_orders': status_stats['total_orders'],
                    'uploaded_orders': status_stats['uploaded_orders'],
                    'sent_orders': status_stats['sent_orders'],
                    'completed_orders': status_stats['completed_orders'],
                    'recent_orders': recent_orders,
                    'overdue_orders': status_stats['overdue_orders'],
                    'factory_stats': list(factory_stats),
                    'monthly_stats': list(monthly_stats),
                    'urgent_orders': list(urgent_orders),
                    'country_stats': list(country_stats)
                }
                cache.set(cache_key, stats, 300)  # 5 минут
                context.update(stats)
            else:
                context.update(cached_stats)
        else:
            # Статистика для неавторизованных пользователей
            cache_key = "public_stats"
            cached_stats = cache.get(cache_key)
            
            if cached_stats is None:
                from django.contrib.auth.models import User
                
                # Общая статистика системы
                public_stats = {
                    'total_orders': Order.objects.count(),
                    'total_factories': Factory.objects.count(),
                    'total_countries': Country.objects.count(),
                    'active_users': User.objects.filter(is_active=True).count(),
                }
                cache.set(cache_key, public_stats, 600)  # 10 минут
                context.update(public_stats)
            else:
                context.update(cached_stats)
        
        return context


class SignUpView(CreateView):
    template_name = 'registration/signup.html'
    form_class = UserCreationForm
    success_url = reverse_lazy('home')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        login(self.request, self.object)
        messages.success(self.request, 'Регистрация прошла успешно!')
        return response


@method_decorator(login_required, name='dispatch')
class OrderListView(ListView):
    model = Order
    template_name = 'orders/order_list.html'
    context_object_name = 'orders'
    paginate_by = 20
    
    def get_queryset(self):
        # Оптимизируем запросы с помощью select_related
        queryset = Order.objects.filter(employee=self.request.user).select_related(
            'factory', 
            'factory__country',
            'employee'
        )
        
        # Фильтрация по статусу
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Фильтрация по фабрике
        factory_id = self.request.GET.get('factory')
        if factory_id:
            queryset = queryset.filter(factory_id=factory_id)
        
        # Поиск
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(factory__name__icontains=search)
            )
        
        return queryset.order_by('-uploaded_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Кэшируем список фабрик на 10 минут
        cache_key = 'active_factories'
        cached_factories = cache.get(cache_key)
        
        if cached_factories is None:
            factories = list(Factory.objects.filter(is_active=True).select_related('country').values(
                'id', 'name', 'country__name', 'country__code'
            ))
            cache.set(cache_key, factories, 600)  # 10 минут
            context['factories'] = factories
        else:
            context['factories'] = cached_factories
            
        context['status_choices'] = Order.STATUS_CHOICES
        return context


@method_decorator(login_required, name='dispatch')
class OrderDetailView(DetailView):
    model = Order
    template_name = 'orders/order_detail.html'
    context_object_name = 'order'
    
    def get_queryset(self):
        return Order.objects.filter(employee=self.request.user).select_related('factory', 'factory__country')


@login_required
def create_order(request):
    if request.method == 'POST':
        form = OrderForm(request.POST, request.FILES)
        if form.is_valid():
            order = form.save(commit=False)
            order.employee = request.user
            order.save()
            
            # Отправляем уведомление о создании заказа
            try:
                send_order_notification.delay(order.id, 'order_uploaded')
            except Exception as e:
                print(f"Ошибка отправки уведомления: {e}")
            
            messages.success(request, 'Заказ успешно создан!')
            return redirect('order_detail', pk=order.pk)
    else:
        form = OrderForm()
    
    return render(request, 'orders/order_form.html', {'form': form})


@login_required
def send_order(request, pk):
    """Создание подтверждения для отправки заказа"""
    order = get_object_or_404(Order, pk=pk, employee=request.user)
    
    if order.status != 'uploaded':
        messages.error(request, 'Заказ уже отправлен или имеет другой статус!')
        return redirect('order_detail', pk=pk)
    
    # Проверяем, есть ли активное подтверждение
    active_confirmation = OrderConfirmation.objects.filter(
        order=order,
        action='send_order',
        status='pending',
        expires_at__gt=timezone.now()
    ).first()
    
    if active_confirmation:
        messages.warning(request, 'У вас уже есть активное подтверждение для отправки этого заказа.')
        return redirect('confirmation_detail', pk=active_confirmation.pk)
    
    # Создаем новое подтверждение
    confirmation = OrderConfirmation.objects.create(
        order=order,
        action='send_order',
        requested_by=request.user,
        confirmation_data={
            'order_title': order.title,
            'factory_name': order.factory.name,
            'factory_email': order.factory.email,
            'country': order.factory.country.name,
        }
    )
    
    messages.info(request, 'Создано подтверждение для отправки заказа. Подтвердите операцию для выполнения.')
    return redirect('confirmation_approve', pk=confirmation.pk)


@login_required
def send_order_execute(request, pk):
    """Выполнение отправки заказа после подтверждения"""
    order = get_object_or_404(Order, pk=pk, employee=request.user)
    
    if order.status != 'uploaded':
        messages.error(request, 'Заказ уже отправлен или имеет другой статус!')
        return redirect('order_detail', pk=pk)
    
    if request.method != 'POST':
        return redirect('send_order', pk=pk)
    
    try:
        # Создаем запись подтверждения
        confirmation = OrderConfirmation.objects.create(
            order=order,
            action='send_order',
            requested_by=request.user,
            confirmation_data={
                'factory_email': order.factory.email,
                'factory_name': order.factory.name,
                'order_title': order.title,
                'order_description': order.description,
                'excel_file_name': order.excel_file.name if order.excel_file else None,
            }
        )
        
        # Определяем язык по стране фабрики
        from .email_utils import get_language_by_country_code, get_email_subject, get_email_template_paths
        
        country_code = order.factory.country.code
        language_code = get_language_by_country_code(country_code)
        
        # Получаем пути к шаблонам
        html_template_path, txt_template_path = get_email_template_paths(language_code)
        
        # Формируем заголовок письма
        subject_prefix = get_email_subject(language_code)
        subject = f'{subject_prefix}: {order.title}'
        
        # Рендерим HTML шаблон
        html_message = render_to_string(html_template_path, {
            'order': order,
            'factory': order.factory,
            'employee': order.employee,
        })
        
        # Рендерим текстовый шаблон
        text_message = render_to_string(txt_template_path, {
            'order': order,
            'factory': order.factory,
            'employee': order.employee,
        })
        
        email = EmailMessage(
            subject=subject,
            body=html_message,  # Используем HTML как основное тело
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[order.factory.email],
        )
        email.content_subtype = "html"  # Указываем HTML контент
        email.encoding = 'utf-8'  # Явно указываем кодировку
        
        # Прикрепляем Excel файл
        if order.excel_file:
            email.attach_file(order.excel_file.path)
        
        email.send()
        
        # Обновляем статус заказа
        order.mark_as_sent()
        
        # Подтверждаем операцию
        confirmation.confirm(request.user, "Заказ успешно отправлен на фабрику")
        
        # Создаем запись аудита
        OrderAuditLog.objects.create(
            order=order,
            action='sent',
            user=request.user,
            old_value='uploaded',
            new_value='sent',
            field_name='status',
            comments=f'Заказ отправлен на фабрику {order.factory.name}'
        )
        
        # Отправляем уведомление о отправке заказа
        try:
            send_order_notification.delay(order.id, 'order_sent')
        except Exception as e:
            print(f"Ошибка отправки уведомления: {e}")
        
        messages.success(request, f'Заказ успешно отправлен на фабрику {order.factory.name}!')
        
    except Exception as e:
        messages.error(request, f'Ошибка при отправке заказа: {str(e)}')
    
    return redirect('order_detail', pk=pk)


@login_required
def upload_invoice(request, pk):
    """Создание подтверждения для загрузки инвойса"""
    order = get_object_or_404(Order, pk=pk, employee=request.user)
    
    # Проверяем, что заказ отправлен (можно прикреплять инвойс только после отправки)
    if order.status not in ['sent', 'invoice_received']:
        messages.error(request, 'Инвойс можно прикрепить только после отправки заказа на фабрику!')
        return redirect('order_detail', pk=pk)
    
    # Проверяем, есть ли активное подтверждение
    active_confirmation = OrderConfirmation.objects.filter(
        order=order,
        action='upload_invoice',
        status='pending',
        expires_at__gt=timezone.now()
    ).first()
    
    if active_confirmation:
        messages.warning(request, 'У вас уже есть активное подтверждение для загрузки инвойса этого заказа.')
        return redirect('confirmation_detail', pk=active_confirmation.pk)
    
    # Создаем новое подтверждение
    confirmation = OrderConfirmation.objects.create(
        order=order,
        action='upload_invoice',
        requested_by=request.user,
        confirmation_data={
            'order_title': order.title,
            'factory_name': order.factory.name,
            'current_status': order.get_status_display(),
            'sent_at': order.sent_at.strftime('%d.%m.%Y %H:%M') if order.sent_at else None,
        }
    )
    
    messages.info(request, 'Создано подтверждение для загрузки инвойса. Подтвердите операцию для выполнения.')
    return redirect('confirmation_approve', pk=confirmation.pk)


@login_required
def upload_invoice_form(request, pk):
    """Показ формы загрузки инвойса после подтверждения"""
    order = get_object_or_404(Order, pk=pk, employee=request.user)
    
    # Проверяем, что заказ отправлен (можно прикреплять инвойс только после отправки)
    if order.status not in ['sent', 'invoice_received']:
        messages.error(request, 'Инвойс можно прикрепить только после отправки заказа на фабрику!')
        return redirect('order_detail', pk=pk)
    
    # Проверяем, есть ли активное подтверждение
    active_confirmation = OrderConfirmation.objects.filter(
        order=order,
        action='upload_invoice',
        status='pending',
        expires_at__gt=timezone.now()
    ).first()
    
    if not active_confirmation:
        messages.error(request, 'Нет активного подтверждения для загрузки инвойса!')
        return redirect('order_detail', pk=pk)
    
    form = InvoiceUploadForm()
    return render(request, 'orders/upload_invoice_form.html', {
        'form': form,
        'order': order,
        'confirmation': active_confirmation
    })


@login_required
def upload_invoice_execute(request, pk):
    """Выполнение загрузки инвойса после подтверждения"""
    order = get_object_or_404(Order, pk=pk, employee=request.user)
    
    # Проверяем, что заказ отправлен (можно прикреплять инвойс только после отправки)
    if order.status not in ['sent', 'invoice_received']:
        messages.error(request, 'Инвойс можно прикрепить только после отправки заказа на фабрику!')
        return redirect('order_detail', pk=pk)
    
    # Проверяем, есть ли активное подтверждение
    active_confirmation = OrderConfirmation.objects.filter(
        order=order,
        action='upload_invoice',
        status='pending',
        expires_at__gt=timezone.now()
    ).first()
    
    if not active_confirmation:
        messages.error(request, 'Нет активного подтверждения для загрузки инвойса!')
        return redirect('order_detail', pk=pk)
    
    if request.method != 'POST':
        return redirect('upload_invoice_form', pk=pk)
    
    form = InvoiceUploadForm(request.POST, request.FILES)
    if form.is_valid():
        try:
            invoice_file = form.cleaned_data['invoice_file']
            old_status = order.status
            
            # Обновляем данные подтверждения
            active_confirmation.confirmation_data.update({
                'invoice_file_name': invoice_file.name,
                'invoice_file_size': invoice_file.size,
            })
            active_confirmation.save()
            
            # Обновляем статус заказа
            order.mark_invoice_received(invoice_file)
            
            # Подтверждаем операцию
            active_confirmation.confirm(request.user, f"Инвойс {invoice_file.name} успешно загружен")
            
            # Создаем запись аудита
            OrderAuditLog.objects.create(
                order=order,
                action='file_uploaded',
                user=request.user,
                old_value=old_status,
                new_value='invoice_received',
                field_name='status',
                comments=f'Загружен инвойс: {invoice_file.name}'
            )
            
            # Отправляем уведомление о получении инвойса
            try:
                send_order_notification.delay(order.id, 'invoice_received')
            except Exception as e:
                print(f"Ошибка отправки уведомления: {e}")
            
            messages.success(request, 'Инвойс успешно прикреплен к заказу!')
            
        except Exception as e:
            messages.error(request, f'Ошибка при загрузке инвойса: {str(e)}')
    else:
        messages.error(request, 'Ошибка в форме загрузки инвойса!')
        return redirect('upload_invoice', pk=pk)
    
    return redirect('order_detail', pk=pk)


@login_required
def download_file(request, pk, file_type):
    order = get_object_or_404(Order, pk=pk, employee=request.user)
    
    if file_type == 'excel' and order.excel_file:
        file_path = order.excel_file.path
        filename = f"{order.title}_order.xlsx"
    elif file_type == 'invoice' and order.invoice_file:
        file_path = order.invoice_file.path
        filename = f"{order.title}_invoice.pdf"
    else:
        messages.error(request, 'Файл не найден!')
        return redirect('order_detail', pk=pk)
    
    if os.path.exists(file_path):
        with open(file_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    else:
        messages.error(request, 'Файл не найден на сервере!')
        return redirect('order_detail', pk=pk)


def get_factories(request):
    """AJAX endpoint для получения фабрик по стране"""
    country_id = request.GET.get('country_id')
    if country_id:
        # Кэшируем фабрики по стране на 30 минут
        cache_key = f'factories_country_{country_id}'
        cached_data = cache.get(cache_key)
        
        if cached_data is None:
            factories = Factory.objects.filter(country_id=country_id, is_active=True)
            data = [{'id': f.id, 'name': f.name} for f in factories]
            cache.set(cache_key, data, 1800)  # 30 минут
            return JsonResponse(data, safe=False)
        else:
            return JsonResponse(cached_data, safe=False)
    return JsonResponse([], safe=False)


@method_decorator(login_required, name='dispatch')
class AnalyticsDashboardView(TemplateView):
    """Дашборд с аналитикой"""
    template_name = 'orders/analytics_dashboard.html'
    paginate_by = 10  # Пагинация для просроченных заказов
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Получаем параметры фильтрации
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        # Конвертируем даты
        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            except ValueError:
                date_from = None
        
        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            except ValueError:
                date_to = None
        
        # Кэшируем данные аналитики на 10 минут
        cache_key = f"analytics_{self.request.user.id}_{date_from}_{date_to}"
        cached_analytics = cache.get(cache_key)
        
        if cached_analytics is None:
            # Получаем данные аналитики
            analytics_data = get_analytics_data(
                user=self.request.user,
                date_from=date_from,
                date_to=date_to
            )
            cache.set(cache_key, analytics_data, 600)  # 10 минут
            context.update(analytics_data)
        else:
            context.update(cached_analytics)
        
        # Добавляем пагинацию для просроченных заказов
        from django.core.paginator import Paginator
        overdue_orders = context.get('overdue_orders', [])
        paginator = Paginator(overdue_orders, self.paginate_by)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context['overdue_orders_page'] = page_obj
        
        return context


@login_required
def analytics_export(request):
    """Экспорт аналитики в Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    # Получаем параметры
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    export_type = request.GET.get('type', 'overview')
    
    # Конвертируем даты
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            date_from = None
    
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            date_to = None
    
    # Получаем данные
    analytics_data = get_analytics_data(
        user=request.user,
        date_from=date_from,
        date_to=date_to
    )
    
    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Аналитика заказов"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Заголовок
    ws['A1'] = f"Аналитика заказов за период {analytics_data['date_range']['from']} - {analytics_data['date_range']['to']}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    
    row = 3
    
    if export_type == 'overview':
        # Общая статистика
        ws['A3'] = "Общая статистика"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        ws.merge_cells('A3:F3')
        row += 1
        
        overview = analytics_data['overview']
        stats_data = [
            ['Показатель', 'Количество'],
            ['Всего заказов', overview['total_orders']],
            ['Загружено', overview['uploaded']],
            ['Отправлено', overview['sent']],
            ['Инвойс получен', overview['invoice_received']],
            ['Завершено', overview['completed']],
            ['Отменено', overview['cancelled']],
        ]
        
        for stat in stats_data:
            ws[f'A{row}'] = stat[0]
            ws[f'B{row}'] = stat[1]
            row += 1
        
        row += 1
        
        # KPI метрики
        ws[f'A{row}'] = "Ключевые показатели"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        kpi = analytics_data['kpi_metrics']
        kpi_data = [
            ['Показатель', 'Значение'],
            ['Процент завершенных заказов', f"{kpi['completion_rate']}%"],
            ['Процент заказов с инвойсами', f"{kpi['invoice_rate']}%"],
            ['Среднее время обработки', f"{kpi['avg_processing_time']} дней"],
            ['Просроченных заказов', kpi['overdue_count']],
        ]
        
        for stat in kpi_data:
            ws[f'A{row}'] = stat[0]
            ws[f'B{row}'] = stat[1]
            row += 1
    
    elif export_type == 'factories':
        # Статистика по фабрикам
        ws['A3'] = "Статистика по фабрикам"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        ws.merge_cells('A3:F3')
        row += 1
        
        headers = ['Фабрика', 'Страна', 'Всего заказов', 'Загружено', 'Отправлено', 'Завершено']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        row += 1
        
        for factory in analytics_data['factory_stats']:
            ws[f'A{row}'] = factory['factory__name']
            ws[f'B{row}'] = factory['factory__country__name']
            ws[f'C{row}'] = factory['total_orders']
            ws[f'D{row}'] = factory['uploaded_orders']
            ws[f'E{row}'] = factory['sent_orders']
            ws[f'F{row}'] = factory['completed_orders']
            row += 1
    
    # Настройка ширины колонок
    for column in ws.columns:
        max_length = 0
        # Пропускаем объединенные ячейки
        if hasattr(column[0], 'column_letter'):
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if hasattr(cell, 'value') and cell.value is not None:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Создаем HTTP ответ
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="analytics_{export_type}_{analytics_data["date_range"]["from"]}.xlsx"'
    
    return response


@login_required
def analytics_api(request):
    """API для получения данных аналитики (для AJAX)"""
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    chart_type = request.GET.get('chart_type', 'overview')
    
    # Конвертируем даты
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            date_from = None
    
    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            date_to = None
    
    # Получаем данные
    analytics_data = get_analytics_data(
        user=request.user,
        date_from=date_from,
        date_to=date_to
    )
    
    # Возвращаем только нужные данные для конкретного типа графика
    if chart_type == 'overview':
        return JsonResponse({
            'overview': analytics_data['overview'],
            'kpi_metrics': analytics_data['kpi_metrics']
        })
    elif chart_type == 'factories':
        return JsonResponse({
            'factory_stats': analytics_data['factory_stats']
        })
    elif chart_type == 'time_series':
        return JsonResponse({
            'time_series': analytics_data['time_series']
        })
    else:
        return JsonResponse(analytics_data)


@csrf_exempt
def clear_messages(request):
    """Очистка сообщений из сессии"""
    if request.method == 'POST':
        # Очищаем все сообщения из сессии
        messages.get_messages(request).used = True
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)


# ===== УПРАВЛЕНИЕ СТРАНАМИ =====

@method_decorator(login_required, name='dispatch')
class CountryListView(ListView):
    """Список стран"""
    model = Country
    template_name = 'orders/country_list.html'
    context_object_name = 'countries'
    paginate_by = 20


@method_decorator(login_required, name='dispatch')
class CountryCreateView(CreateView):
    """Создание новой страны"""
    model = Country
    form_class = CountryForm
    template_name = 'orders/country_form.html'
    success_url = reverse_lazy('country_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Страна успешно создана!')
        return super().form_valid(form)


@method_decorator(login_required, name='dispatch')
class CountryUpdateView(UpdateView):
    """Редактирование страны"""
    model = Country
    form_class = CountryForm
    template_name = 'orders/country_form.html'
    success_url = reverse_lazy('country_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Страна успешно обновлена!')
        return super().form_valid(form)


@login_required
def country_delete(request, pk):
    """Удаление страны"""
    country = get_object_or_404(Country, pk=pk)
    
    # Проверяем, есть ли фабрики в этой стране
    factories_count = Factory.objects.filter(country=country).count()
    if factories_count > 0:
        messages.error(request, f'Нельзя удалить страну, в которой есть фабрики ({factories_count} шт.)')
        return redirect('country_list')
    
    if request.method == 'POST':
        country.delete()
        messages.success(request, 'Страна успешно удалена!')
        return redirect('country_list')
    
    return render(request, 'orders/country_confirm_delete.html', {'country': country})


# ===== УПРАВЛЕНИЕ ФАБРИКАМИ =====

@method_decorator(login_required, name='dispatch')
class FactoryListView(ListView):
    """Список фабрик"""
    model = Factory
    template_name = 'orders/factory_list.html'
    context_object_name = 'factories'
    paginate_by = 20
    
    def get_queryset(self):
        return Factory.objects.select_related('country').order_by('country__name', 'name')


@method_decorator(login_required, name='dispatch')
class FactoryCreateView(CreateView):
    """Создание новой фабрики"""
    model = Factory
    form_class = FactoryForm
    template_name = 'orders/factory_form.html'
    success_url = reverse_lazy('factory_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Фабрика успешно создана!')
        return super().form_valid(form)


@method_decorator(login_required, name='dispatch')
class FactoryUpdateView(UpdateView):
    """Редактирование фабрики"""
    model = Factory
    form_class = FactoryForm
    template_name = 'orders/factory_form.html'
    success_url = reverse_lazy('factory_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Фабрика успешно обновлена!')
        return super().form_valid(form)


@login_required
def factory_delete(request, pk):
    """Удаление фабрики"""
    factory = get_object_or_404(Factory, pk=pk)
    
    # Проверяем, есть ли заказы для этой фабрики
    orders_count = Order.objects.filter(factory=factory).count()
    if orders_count > 0:
        messages.error(request, f'Нельзя удалить фабрику, для которой есть заказы ({orders_count} шт.)')
        return redirect('factory_list')
    
    if request.method == 'POST':
        factory.delete()
        messages.success(request, 'Фабрика успешно удалена!')
        return redirect('factory_list')
    
    return render(request, 'orders/factory_confirm_delete.html', {'factory': factory})


# ===== AJAX ENDPOINTS =====

@login_required
def get_countries(request):
    """AJAX endpoint для получения списка стран"""
    countries = Country.objects.all().order_by('name')
    data = [{'id': country.id, 'name': country.name} for country in countries]
    return JsonResponse({'countries': data})


@login_required
def create_country_ajax(request):
    """AJAX endpoint для создания страны"""
    if request.method == 'POST':
        form = CountryForm(request.POST)
        if form.is_valid():
            country = form.save()
            return JsonResponse({
                'success': True,
                'country': {
                    'id': country.id,
                    'name': country.name,
                    'code': country.code
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def create_factory_ajax(request):
    """AJAX endpoint для создания фабрики"""
    if request.method == 'POST':
        form = FactoryForm(request.POST)
        if form.is_valid():
            factory = form.save()
            return JsonResponse({
                'success': True,
                'factory': {
                    'id': factory.id,
                    'name': factory.name,
                    'country': factory.country.name,
                    'country_id': factory.country.id
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

# ===== ПРЕДВАРИТЕЛЬНЫЙ ПРОСМОТР ФАЙЛОВ =====

@login_required
def preview_file(request, pk, file_type):
    """Предварительный просмотр файла"""
    order = get_object_or_404(Order, pk=pk, employee=request.user)
    
    # Определяем путь к файлу
    if file_type == 'excel' and order.excel_file:
        file_path = order.excel_file.path
    elif file_type == 'invoice' and order.invoice_file:
        file_path = order.invoice_file.path
    else:
        return JsonResponse({'error': 'Файл не найден'}, status=404)
    
    # Генерируем предварительный просмотр
    # Для invoice файлов используем тип 'pdf'
    preview_type = 'pdf' if file_type == 'invoice' else file_type
    preview_data = generate_file_preview(file_path, preview_type)
    
    return JsonResponse(preview_data)


@login_required
def preview_file_modal(request, pk, file_type):
    """Модальное окно предварительного просмотра файла"""
    order = get_object_or_404(Order, pk=pk, employee=request.user)
    
    # Определяем файл
    if file_type == 'excel' and order.excel_file:
        file = order.excel_file
        file_name = file.name
    elif file_type == 'invoice' and order.invoice_file:
        file = order.invoice_file
        file_name = file.name
    else:
        messages.error(request, 'Файл не найден')
        return redirect('order_detail', pk=pk)
    
    return render(request, 'orders/file_preview_modal.html', {
        'order': order,
        'file': file,
        'file_type': file_type,
        'file_name': file_name
    })


@login_required
def notification_settings(request):
    """Настройки уведомлений пользователя"""
    try:
        settings_obj = NotificationSettings.objects.get(user=request.user)
    except NotificationSettings.DoesNotExist:
        settings_obj = NotificationSettings.objects.create(user=request.user)
    
    if request.method == 'POST':
        form = NotificationSettingsForm(request.POST, instance=settings_obj)
        if form.is_valid():
            form.save()
            messages.success(request, 'Настройки уведомлений сохранены!')
            return redirect('notification_settings')
    else:
        form = NotificationSettingsForm(instance=settings_obj)
    
    return render(request, 'orders/notification_settings.html', {
        'form': form,
        'settings': settings_obj
    })


@login_required
def test_notification(request):
    """Тестирование отправки уведомления"""
    if request.method == 'POST':
        # Получаем последний заказ пользователя
        last_order = Order.objects.filter(employee=request.user).first()
        
        if last_order:
            # Отправляем тестовое уведомление
            send_order_notification.delay(last_order.id, 'order_sent')
            messages.success(request, 'Тестовое уведомление отправлено! Проверьте email.')
        else:
            messages.warning(request, 'У вас нет заказов для тестирования.')
        
        return redirect('notification_settings')
    
    return redirect('notification_settings')


@login_required
def complete_order(request, pk):
    """Создание подтверждения для завершения заказа"""
    order = get_object_or_404(Order, pk=pk, employee=request.user)
    
    # Проверяем, что заказ можно завершить
    if order.status not in ['invoice_received']:
        messages.error(request, 'Заказ можно завершить только после получения инвойса!')
        return redirect('order_detail', pk=pk)
    
    # Проверяем, есть ли активное подтверждение
    active_confirmation = OrderConfirmation.objects.filter(
        order=order,
        action='complete_order',
        status='pending',
        expires_at__gt=timezone.now()
    ).first()
    
    if active_confirmation:
        messages.warning(request, 'У вас уже есть активное подтверждение для завершения этого заказа.')
        return redirect('confirmation_detail', pk=active_confirmation.pk)
    
    # Создаем новое подтверждение
    confirmation = OrderConfirmation.objects.create(
        order=order,
        action='complete_order',
        requested_by=request.user,
        confirmation_data={
            'order_title': order.title,
            'factory_name': order.factory.name,
            'current_status': order.get_status_display(),
            'uploaded_at': order.uploaded_at.strftime('%d.%m.%Y %H:%M'),
            'sent_at': order.sent_at.strftime('%d.%m.%Y %H:%M') if order.sent_at else None,
            'invoice_received_at': order.invoice_received_at.strftime('%d.%m.%Y %H:%M') if order.invoice_received_at else None,
        }
    )
    
    messages.info(request, 'Создано подтверждение для завершения заказа. Подтвердите операцию для выполнения.')
    return redirect('confirmation_approve', pk=confirmation.pk)


# ===== УПРАВЛЕНИЕ ПОДТВЕРЖДЕНИЯМИ =====

@login_required
def confirmation_list(request):
    """Список подтверждений пользователя"""
    confirmations = OrderConfirmation.objects.filter(
        order__employee=request.user
    ).select_related('order', 'requested_by', 'confirmed_by').order_by('-requested_at')
    
    # Фильтрация по статусу
    status_filter = request.GET.get('status')
    if status_filter:
        confirmations = confirmations.filter(status=status_filter)
    
    # Пагинация
    from django.core.paginator import Paginator
    paginator = Paginator(confirmations, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'orders/confirmation_list.html', {
        'page_obj': page_obj,
        'status_choices': OrderConfirmation._meta.get_field('status').choices,
        'status_filter': status_filter,
    })


@login_required
def confirmation_detail(request, pk):
    """Детальная страница подтверждения"""
    confirmation = get_object_or_404(OrderConfirmation, pk=pk, order__employee=request.user)
    
    return render(request, 'orders/confirmation_detail.html', {
        'confirmation': confirmation,
    })


@login_required
def confirmation_approve(request, pk):
    """Подтверждение операции"""
    confirmation = get_object_or_404(OrderConfirmation, pk=pk, order__employee=request.user)
    
    if confirmation.status != 'pending':
        messages.error(request, 'Это подтверждение уже обработано!')
        return redirect('confirmation_detail', pk=pk)
    
    if confirmation.is_expired():
        messages.error(request, 'Срок подтверждения истек!')
        return redirect('confirmation_detail', pk=pk)
    
    if request.method == 'POST':
        comments = request.POST.get('comments', '')
        
        try:
            # Выполняем соответствующее действие
            if confirmation.action == 'send_order':
                # Выполняем отправку заказа
                order = confirmation.order
                
                # Проверяем, что заказ еще не отправлен
                if order.status != 'uploaded':
                    messages.error(request, 'Заказ уже отправлен или имеет другой статус!')
                    return redirect('confirmation_detail', pk=pk)
                
                # Отправляем email фабрике
                try:
                    # Определяем язык по стране фабрики
                    from .email_utils import get_language_by_country_code, get_email_subject, get_email_template_paths
                    
                    country_code = order.factory.country.code
                    language_code = get_language_by_country_code(country_code)
                    
                    # Получаем пути к шаблонам
                    html_template_path, txt_template_path = get_email_template_paths(language_code)
                    
                    # Формируем заголовок письма
                    subject_prefix = get_email_subject(language_code)
                    subject = f'{subject_prefix}: {order.title}'
                    
                    # Рендерим HTML шаблон
                    html_message = render_to_string(html_template_path, {
                        'order': order,
                        'factory': order.factory,
                        'employee': order.employee,
                    })
                    
                    # Рендерим текстовый шаблон
                    text_message = render_to_string(txt_template_path, {
                        'order': order,
                        'factory': order.factory,
                        'employee': order.employee,
                    })
                    
                    email = EmailMessage(
                        subject=subject,
                        body=html_message,  # Используем HTML как основное тело
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[order.factory.email],
                    )
                    email.content_subtype = "html"  # Указываем HTML контент
                    email.encoding = 'utf-8'  # Явно указываем кодировку
                    
                    # Прикрепляем Excel файл
                    if order.excel_file:
                        email.attach_file(order.excel_file.path)
                    
                    email.send()
                    
                except Exception as e:
                    print(f"Ошибка отправки email: {e}")
                    messages.error(request, f'Ошибка при отправке email: {str(e)}')
                    return redirect('confirmation_detail', pk=pk)
                
                # Обновляем статус заказа
                order.status = 'sent'
                order.sent_at = timezone.now()
                order.save()
                
                # Создаем запись аудита
                OrderAuditLog.objects.create(
                    order=order,
                    action='sent',
                    user=request.user,
                    old_value='uploaded',
                    new_value='sent',
                    field_name='status',
                    comments='Заказ отправлен на фабрику'
                )
                
                # Отправляем уведомление
                try:
                    send_order_notification.delay(order.id, 'order_sent')
                except Exception as e:
                    print(f"Ошибка отправки уведомления: {e}")
                
                messages.success(request, f'Заказ успешно отправлен на фабрику {order.factory.name}!')
                
            elif confirmation.action == 'upload_invoice':
                # Для загрузки инвойса нужна форма с файлом
                # Перенаправляем на специальную страницу загрузки инвойса
                return redirect('upload_invoice_form', pk=confirmation.order.pk)
                
            elif confirmation.action == 'complete_order':
                # Выполняем завершение заказа
                order = confirmation.order
                
                # Проверяем, что заказ можно завершить
                if order.status not in ['invoice_received']:
                    messages.error(request, 'Заказ можно завершить только после получения инвойса!')
                    return redirect('confirmation_detail', pk=pk)
                
                # Завершаем заказ
                order.mark_as_completed()
                
                # Создаем запись аудита
                OrderAuditLog.objects.create(
                    order=order,
                    action='completed',
                    user=request.user,
                    old_value='invoice_received',
                    new_value='completed',
                    field_name='status',
                    comments='Заказ завершен пользователем'
                )
                
                # Отправляем уведомление о завершении
                try:
                    send_order_notification.delay(order.id, 'order_completed')
                except Exception as e:
                    print(f"Ошибка отправки уведомления: {e}")
                
                messages.success(request, f'Заказ "{order.title}" успешно завершен!')
            
            # Подтверждаем операцию
            confirmation.confirm(request.user, comments)
            
        except Exception as e:
            messages.error(request, f'Ошибка при подтверждении: {str(e)}')
        
        return redirect('confirmation_detail', pk=pk)
    
    return render(request, 'orders/confirmation_approve.html', {
        'confirmation': confirmation,
    })


@login_required
def confirmation_reject(request, pk):
    """Отклонение операции"""
    confirmation = get_object_or_404(OrderConfirmation, pk=pk, order__employee=request.user)
    
    if confirmation.status != 'pending':
        messages.error(request, 'Это подтверждение уже обработано!')
        return redirect('confirmation_detail', pk=pk)
    
    if request.method == 'POST':
        reason = request.POST.get('reason', '')
        
        try:
            confirmation.reject(request.user, reason)
            messages.success(request, 'Операция отклонена!')
            
        except Exception as e:
            messages.error(request, f'Ошибка при отклонении: {str(e)}')
        
        return redirect('confirmation_detail', pk=pk)
    
    return render(request, 'orders/confirmation_reject.html', {
        'confirmation': confirmation,
    })


@login_required
def notification_list(request):
    """Список уведомлений пользователя"""
    from django.core.paginator import Paginator
    
    # Получаем все уведомления пользователя
    notifications = Notification.objects.filter(user=request.user)
    
    # Создаем форму фильтрации
    form = NotificationFilterForm(request.GET)
    
    # Применяем фильтры только если форма валидна
    if form.is_valid():
        search = form.cleaned_data.get('search')
        notification_type = form.cleaned_data.get('notification_type')
        status = form.cleaned_data.get('status')
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        
        # Поиск по тексту
        if search:
            from django.db.models import Q
            notifications = notifications.filter(
                Q(title__icontains=search) | Q(message__icontains=search)
            )
        
        # Фильтр по типу
        if notification_type:
            notifications = notifications.filter(notification_type=notification_type)
        
        # Фильтр по статусу
        if status == 'unread':
            notifications = notifications.filter(is_read=False)
        elif status == 'read':
            notifications = notifications.filter(is_read=True)
        
        # Фильтр по дате
        if date_from:
            notifications = notifications.filter(created_at__date__gte=date_from)
        if date_to:
            notifications = notifications.filter(created_at__date__lte=date_to)
    
    # Сортировка
    notifications = notifications.order_by('-created_at')
    
    # Пагинация
    paginator = Paginator(notifications, 15)  # 15 уведомлений на страницу
    page_number = request.GET.get('page')
    notifications_page = paginator.get_page(page_number)
    
    # Уведомления помечаются как прочитанные только при явном действии пользователя
    # Убрали автоматическую пометку при просмотре списка
    
    return render(request, 'orders/notification_list.html', {
        'notifications': notifications_page,
        'form': form,
        'total_count': notifications.count(),
        'unread_count': notifications.filter(is_read=False).count()
    })


@login_required
def mark_notification_read(request, pk):
    """Пометить уведомление как прочитанное"""
    notification = get_object_or_404(Notification, pk=pk, user=request.user)
    notification.mark_as_read()
    
    return JsonResponse({'status': 'success'})


@login_required
def mark_all_notifications_read(request):
    """Пометить все уведомления как прочитанные"""
    if request.method == 'POST':
        unread_notifications = Notification.objects.filter(user=request.user, is_read=False)
        count = unread_notifications.count()
        unread_notifications.update(is_read=True, read_at=timezone.now())
        
        return JsonResponse({'status': 'success', 'count': count})
    
    return JsonResponse({'status': 'error'}, status=400)
