import os
import json
from typing import Dict, List, Any, Optional
from django.conf import settings
from django.core.files.storage import default_storage
from django.http import JsonResponse, HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
import xlrd
import io
import base64
from PIL import Image
from pdf2image import convert_from_path


class FilePreviewGenerator:
    """Генератор предварительного просмотра файлов"""
    
    @staticmethod
    def preview_excel(file_path: str, max_rows: int = 10) -> Dict[str, Any]:
        """Предварительный просмотр Excel файла"""
        try:
            file_extension = os.path.splitext(file_path)[1].lower()
            
            # Собираем данные
            data = {
                'type': 'excel',
                'filename': os.path.basename(file_path),
                'sheets': [],
                'preview_data': []
            }
            
            if file_extension == '.xlsx':
                # Используем openpyxl для .xlsx файлов
                workbook = openpyxl.load_workbook(file_path, read_only=True)
                worksheet = workbook.active
                
                # Информация о листах
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    data['sheets'].append({
                        'name': sheet_name,
                        'max_row': sheet.max_row,
                        'max_column': sheet.max_column
                    })
                
                # Предварительный просмотр данных (первые строки)
                preview_rows = []
                for row_num in range(1, min(max_rows + 1, worksheet.max_row + 1)):
                    row_data = []
                    for col_num in range(1, min(11, worksheet.max_column + 1)):  # Максимум 10 колонок
                        cell = worksheet.cell(row=row_num, column=col_num)
                        row_data.append({
                            'value': str(cell.value) if cell.value is not None else '',
                            'column': get_column_letter(col_num)
                        })
                    preview_rows.append(row_data)
                
                data['preview_data'] = preview_rows
                data['total_rows'] = worksheet.max_row
                data['total_columns'] = worksheet.max_column
                
                workbook.close()
                
            elif file_extension == '.xls':
                # Используем xlrd для .xls файлов
                workbook = xlrd.open_workbook(file_path)
                worksheet = workbook.sheet_by_index(0)  # Первый лист
                
                # Информация о листах
                for i in range(workbook.nsheets):
                    sheet = workbook.sheet_by_index(i)
                    data['sheets'].append({
                        'name': sheet.name,
                        'max_row': sheet.nrows,
                        'max_column': sheet.ncols
                    })
                
                # Предварительный просмотр данных (первые строки)
                preview_rows = []
                for row_num in range(min(max_rows, worksheet.nrows)):
                    row_data = []
                    for col_num in range(min(10, worksheet.ncols)):  # Максимум 10 колонок
                        cell_value = worksheet.cell_value(row_num, col_num)
                        # Конвертируем в строку и обрабатываем типы
                        if isinstance(cell_value, float):
                            # Проверяем, является ли это датой
                            if xlrd.xldate.xldate_as_datetime(cell_value, workbook.datemode):
                                cell_value = xlrd.xldate.xldate_as_datetime(cell_value, workbook.datemode).strftime('%Y-%m-%d %H:%M:%S')
                            else:
                                cell_value = str(int(cell_value)) if cell_value.is_integer() else str(cell_value)
                        else:
                            cell_value = str(cell_value) if cell_value is not None else ''
                        
                        row_data.append({
                            'value': cell_value,
                            'column': xlrd.colname(col_num)
                        })
                    preview_rows.append(row_data)
                
                data['preview_data'] = preview_rows
                data['total_rows'] = worksheet.nrows
                data['total_columns'] = worksheet.ncols
            
            return data
            
        except Exception as e:
            return {
                'type': 'excel',
                'filename': os.path.basename(file_path),
                'error': f'Ошибка чтения Excel файла: {str(e)}'
            }
    
    @staticmethod
    def preview_pdf(file_path: str, max_pages: int = 3) -> Dict[str, Any]:
        """Предварительный просмотр PDF файла"""
        try:
            # Конвертируем PDF в изображения
            images = convert_from_path(file_path, first_page=1, last_page=max_pages, dpi=150)
            
            data = {
                'type': 'pdf',
                'filename': os.path.basename(file_path),
                'total_pages': len(images),
                'preview_pages': []
            }
            
            # Генерируем превью страниц
            for page_num, image in enumerate(images):
                # Конвертируем в base64
                img_buffer = io.BytesIO()
                image.save(img_buffer, format='PNG')
                img_data = img_buffer.getvalue()
                img_base64 = base64.b64encode(img_data).decode()
                
                data['preview_pages'].append({
                    'page_number': page_num + 1,
                    'image': f'data:image/png;base64,{img_base64}',
                    'width': image.width,
                    'height': image.height
                })
            
            return data
            
        except Exception as e:
            return {
                'type': 'pdf',
                'filename': os.path.basename(file_path),
                'error': f'Ошибка чтения PDF файла: {str(e)}'
            }
    
    @staticmethod
    def get_file_info(file_path: str) -> Dict[str, Any]:
        """Получение информации о файле"""
        try:
            stat = os.stat(file_path)
            return {
                'size': stat.st_size,
                'size_mb': round(stat.st_size / (1024 * 1024), 2),
                'modified': stat.st_mtime,
                'extension': os.path.splitext(file_path)[1].lower()
            }
        except Exception as e:
            return {'error': f'Ошибка получения информации о файле: {str(e)}'}


def generate_file_preview(file_path: str, file_type: str) -> Dict[str, Any]:
    """Генерация предварительного просмотра файла"""
    if not os.path.exists(file_path):
        return {'error': 'Файл не найден'}
    
    # Получаем информацию о файле
    file_info = FilePreviewGenerator.get_file_info(file_path)
    
    if 'error' in file_info:
        return file_info
    
    # Определяем тип файла по расширению
    file_extension = os.path.splitext(file_path)[1].lower()
    
    # Генерируем предварительный просмотр в зависимости от расширения файла
    if file_extension in ['.xlsx', '.xls']:
        preview = FilePreviewGenerator.preview_excel(file_path)
    elif file_extension == '.pdf':
        preview = FilePreviewGenerator.preview_pdf(file_path)
    elif file_extension in ['.doc', '.docx']:
        # Для Word файлов показываем только информацию о файле
        preview = {
            'type': 'word',
            'filename': os.path.basename(file_path),
            'message': 'Предварительный просмотр Word файлов не поддерживается. Скачайте файл для просмотра.'
        }
    else:
        return {'error': f'Неподдерживаемый тип файла: {file_extension}'}
    
    # Добавляем информацию о файле
    preview.update(file_info)
    
    return preview
